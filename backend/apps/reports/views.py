"""apps/reports/views.py – Dashboard summary + all report endpoints."""
from decimal import Decimal
from datetime import date, timedelta
import io

from django.db.models import Sum, Q
from django.utils import timezone
from django.http import HttpResponse

from rest_framework.views import APIView
from rest_framework.response import Response

from apps.trucks.models import Truck
from apps.drivers.models import Driver
from apps.trips.models import Trip
from apps.fuel.models import FuelLog
from apps.finance.models import Revenue, Expenditure
from apps.inventory.models import Item, StockLedger
from apps.invoicing.models import Invoice
from apps.maintenance.models import MaintenanceLog
from apps.tyres.models import Tyre, TyreAssignment


class BranchFilterMixin:
    def _bf(self):
        return _branch_filter(self.request)


def _parse_dates(request):
    today = timezone.now().date()
    try:
        date_from = date.fromisoformat(request.query_params.get('date_from', ''))
    except (ValueError, TypeError):
        date_from = today.replace(day=1)
    try:
        date_to = date.fromisoformat(request.query_params.get('date_to', ''))
    except (ValueError, TypeError):
        date_to = today
    return date_from, date_to


def _fmt(val):
    if val is None:
        return 0
    return float(round(Decimal(str(val)), 2))


def _branch_filter(request):
    """Returns a dict filter for branch scoping.
    - SUPER_ADMIN: uses ?branch_id= param if given, else no filter (all branches).
    - All others: scoped to their assigned branch.
    """
    user = request.user
    if getattr(user, 'role', None) == 'SUPER_ADMIN':
        param = request.query_params.get('branch_id')
        if param:
            try:
                return {'branch_id': int(param)}
            except (ValueError, TypeError):
                pass
        return {}  # SUPER_ADMIN with no filter = all branches
    if user.branch_id:
        return {'branch_id': user.branch_id}
    return {'branch_id': -1}


def _apply_branch(qs, request):
    """Applies branch filter to a queryset.
    - SUPER_ADMIN: uses ?branch_id= param if given, else returns full qs.
    - All others: scoped to their assigned branch.
    """
    user = request.user
    if getattr(user, 'role', None) == 'SUPER_ADMIN':
        param = request.query_params.get('branch_id')
        if param:
            try:
                return qs.filter(branch_id=int(param))
            except (ValueError, TypeError):
                pass
        return qs  # all branches
    if user.branch_id:
        return qs.filter(branch_id=user.branch_id)
    return qs.none()


def _export_excel(headers, rows, sheet_name='Report'):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='1a56db')
            cell.alignment = Alignment(horizontal='center')
        for row in rows:
            ws.append(list(row))
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        resp['Content-Disposition'] = f'attachment; filename="{sheet_name}.xlsx"'
        return resp
    except ImportError:
        return HttpResponse('openpyxl not installed', status=500)


def _export_pdf(headers, rows, title='Report'):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                leftMargin=20, rightMargin=20,
                                topMargin=30, bottomMargin=30)
        styles = getSampleStyleSheet()
        elements = [Paragraph(title, styles['Title'])]
        data = [headers] + [list(r) for r in rows]
        tbl = Table(data, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a56db')),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 0), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ]))
        elements.append(tbl)
        doc.build(elements)
        buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="{title}.pdf"'
        return resp
    except ImportError:
        return HttpResponse('reportlab not installed', status=500)


def _respond(request, headers, rows, summary=None, sheet_name='Report'):
    export = request.query_params.get('export', 'json')
    if export == 'excel':
        return _export_excel(headers, rows, sheet_name)
    if export == 'pdf':
        return _export_pdf(headers, rows, sheet_name)
    return Response({'headers': headers, 'rows': rows, 'summary': summary or {}})


class DashboardSummaryView(BranchFilterMixin, APIView):
    def get(self, request):
        import logging
        _logger = logging.getLogger('apps.reports')
        try:
            today       = timezone.now().date()
            month_start = today.replace(day=1)

            # SUPER_ADMIN: respect ?branch_id= filter or show ALL branches
            if request.user.role == 'SUPER_ADMIN':
                param_branch = request.query_params.get('branch_id')
                if param_branch:
                    try:
                        branch_id = int(param_branch)
                    except (ValueError, TypeError):
                        branch_id = request.user.branch_id
                else:
                    branch_id = None  # None = all branches for SUPER_ADMIN
            else:
                branch_id = request.user.branch_id

            def bfilter(qs):
                if branch_id is not None:
                    return qs.filter(branch_id=branch_id)
                # SUPER_ADMIN viewing all branches — no branch filter
                if request.user.role == 'SUPER_ADMIN':
                    return qs
                return qs.none()

            active_trucks  = bfilter(Truck.objects.filter(status=Truck.ACTIVE)).count()
            active_drivers = bfilter(Driver.objects.filter(status=Driver.ACTIVE)).count()
            ongoing_trips  = bfilter(Trip.objects.filter(
                status__in=[Trip.EN_ROUTE, Trip.PLANNED]
            )).count()
            trips_this_month = bfilter(Trip.objects.filter(
                loading_time__date__gte=month_start,
                loading_time__date__lte=today,
            )).count()
            monthly_revenue = bfilter(Revenue.objects.filter(
                date__gte=month_start, date__lte=today
            )).aggregate(total=Sum('amount'))['total'] or Decimal('0')
            monthly_expenditure = bfilter(Expenditure.objects.filter(
                date__gte=month_start, date__lte=today
            )).aggregate(total=Sum('amount'))['total'] or Decimal('0')
            _fuel_qs = bfilter(FuelLog.objects.filter(date__gte=month_start, date__lte=today))
            fuel_litres        = _fuel_qs.aggregate(t=Sum('litres'))['t'] or Decimal('0')
            fuel_excess_events = _fuel_qs.filter(excess_fuel__gt=0).count()
            stock_qs = StockLedger.objects.all()
            if branch_id is not None:
                stock_qs = stock_qs.filter(branch_id=branch_id)
            elif request.user.role != 'SUPER_ADMIN':
                stock_qs = StockLedger.objects.none()
            stock_value = stock_qs.aggregate(total=Sum('final_amount'))['total'] or Decimal('0')
            items_qs = Item.objects.filter(ledger_entries__branch_id=branch_id).distinct() if branch_id is not None else (
                Item.objects.filter(ledger_entries__isnull=False).distinct() if request.user.role == 'SUPER_ADMIN'
                else Item.objects.none()
            )
            stock_items = items_qs.count()

            alerts = []
            for truck in bfilter(Truck.objects.filter(status=Truck.ACTIVE)):
                try:
                    alerts.extend(truck.expiry_alerts())
                except Exception:
                    pass
            alerts.sort(key=lambda a: a['days_remaining'])

            try:
                breakdown = self._truck_breakdown(month_start, today, branch_id)
            except Exception as e:
                _logger.error('Dashboard truck_breakdown failed: %s', e, exc_info=True)
                breakdown = []

            return Response({
                'fleet': {
                    'active_trucks':  active_trucks,
                    'active_drivers': active_drivers,
                    'ongoing_trips':  ongoing_trips,
                },
                'this_month': {
                    'trips':              trips_this_month,
                    'revenue':            str(_fmt(monthly_revenue)),
                    'expenditure':        str(_fmt(monthly_expenditure)),
                    'fuel_litres':        _fmt(fuel_litres),
                    'fuel_excess_events': fuel_excess_events,
                },
                'stock_value': _fmt(stock_value),
                'stock_items': stock_items,
                'expiry_alerts':   alerts,
                'truck_breakdown': breakdown,
            })

        except Exception as e:
            _logger.error('Dashboard summary failed: %s', e, exc_info=True)
            return Response(
                {'detail': f'Dashboard data failed to load: {str(e)}'},
                status=500
            )

    def _truck_breakdown(self, date_from, date_to, branch_id=None):
        rows = []
        qs = Truck.objects.filter(status=Truck.ACTIVE)
        if branch_id is not None:
            qs = qs.filter(branch_id=branch_id)
        # if branch_id is None = SUPER_ADMIN viewing all branches, no filter applied
        for truck in qs:
            trips = Trip.objects.filter(
                truck=truck,
                loading_time__date__gte=date_from,
                loading_time__date__lte=date_to,
            )
            trip_rev  = _fmt(trips.aggregate(t=Sum('trip_revenue'))['t'])
            total_exp = _fmt(Expenditure.objects.filter(
                truck=truck, date__gte=date_from, date__lte=date_to
            ).aggregate(t=Sum('amount'))['t'])
            net = round(trip_rev - total_exp, 2)
            rows.append({
                'truck':       truck.truck_number,
                'model':       truck.model,
                'trips':       trips.count(),
                'revenue':     trip_rev,
                'expenditure': total_exp,
                'net':         net,
            })
        rows.sort(key=lambda r: -r['revenue'])
        return rows


class RevenueExpenditureReportView(BranchFilterMixin, APIView):
    def get(self, request):
        date_from, date_to = _parse_dates(request)
        revenues     = _apply_branch(Revenue.objects.filter(date__gte=date_from, date__lte=date_to), request)
        expenditures = _apply_branch(Expenditure.objects.filter(date__gte=date_from, date__lte=date_to), request)
        total_rev = revenues.aggregate(t=Sum('amount'))['t'] or Decimal('0')
        total_exp = expenditures.aggregate(t=Sum('amount'))['t'] or Decimal('0')
        net       = total_rev - total_exp
        headers = ['Date', 'Type', 'Category / Source', 'Truck', 'Description', 'Reference', 'Amount (GH₵)']
        rows = []
        for r in revenues.order_by('date'):
            truck_num = r.trip.truck.truck_number if r.trip_id and r.trip else '—'
            rows.append([str(r.date), 'Revenue', r.get_source_display(),
                         truck_num, r.description or '', r.reference or '', _fmt(r.amount)])
        for e in expenditures.order_by('date'):
            truck_num = e.truck.truck_number if e.truck_id else '—'
            rows.append([str(e.date), 'Expenditure', e.get_category_display(),
                         truck_num, e.description or '', e.reference or '', _fmt(e.amount)])
        rows.sort(key=lambda x: x[0])
        summary = {
            'Total Revenue':     _fmt(total_rev),
            'Total Expenditure': _fmt(total_exp),
            'Net Profit / Loss': _fmt(net),
        }
        return _respond(request, headers, rows, summary, 'Revenue vs Expenditure')


class FuelReportView(BranchFilterMixin, APIView):
    def get(self, request):
        date_from, date_to = _parse_dates(request)
        logs = _apply_branch(FuelLog.objects.filter(
            date__gte=date_from, date__lte=date_to
        ), request).select_related('truck', 'trip').order_by('date')
        headers = ['Date', 'Truck', 'Trip / Waybill', 'Litres', 'Limit', 'Excess',
                   'Price/L (GH₵)', 'Total Cost (GH₵)', 'Remark']
        rows = []
        for fl in logs:
            rows.append([
                str(fl.date), fl.truck.truck_number,
                fl.trip.waybill_no if fl.trip else '—',
                _fmt(fl.litres), _fmt(fl.fuel_limit), _fmt(fl.excess_fuel),
                _fmt(fl.price_per_litre), _fmt(fl.total_cost), fl.remark or '',
            ])
        agg = logs.aggregate(
            total_litres=Sum('litres'), total_excess=Sum('excess_fuel'), total_cost=Sum('total_cost'),
        )
        summary = {
            'Total Litres':     _fmt(agg['total_litres']),
            'Total Excess (L)': _fmt(agg['total_excess']),
            'Total Cost (GH₵)': _fmt(agg['total_cost']),
            'Excess Events':    logs.filter(excess_fuel__gt=0).count(),
        }
        return _respond(request, headers, rows, summary, 'Fuel Report')


class TripReportView(BranchFilterMixin, APIView):
    def get(self, request):
        date_from, date_to = _parse_dates(request)
        truck_id = request.query_params.get('truck')
        qs = _apply_branch(Trip.objects.filter(
            loading_time__date__gte=date_from,
            loading_time__date__lte=date_to,
        ), request).select_related('truck', 'driver')
        if truck_id:
            qs = qs.filter(truck_id=truck_id)
        qs = qs.order_by('loading_time')
        headers = ['Waybill', 'Truck', 'Driver', 'Origin', 'Destination', 'Material',
                   'Loaded (t)', 'Delivered (t)', 'Status', 'Revenue (GH₵)', 'Loading Date']
        rows = []
        for t in qs:
            rows.append([
                t.waybill_no, t.truck.truck_number, t.driver.name,
                t.origin, t.destination, t.material_type,
                _fmt(t.loaded_qty),
                _fmt(t.delivered_qty) if t.delivered_qty is not None else '—',
                t.get_status_display(), _fmt(t.trip_revenue),
                str(t.loading_time.date()),
            ])
        agg = qs.aggregate(
            total_rev=Sum('trip_revenue'),
            total_loaded=Sum('loaded_qty'),
            total_delivered=Sum('delivered_qty'),
        )
        summary = {
            'Total Trips':         qs.count(),
            'Total Revenue (GH₵)': _fmt(agg['total_rev']),
            'Total Loaded (t)':    _fmt(agg['total_loaded']),
            'Total Delivered (t)': _fmt(agg['total_delivered']),
        }
        return _respond(request, headers, rows, summary, 'Trip Report')


class TripDetailReportView(BranchFilterMixin, APIView):
    def get(self, request):
        date_from, date_to = _parse_dates(request)
        truck_id = request.query_params.get('truck')
        qs = _apply_branch(Trip.objects.filter(
            loading_time__date__gte=date_from,
            loading_time__date__lte=date_to,
        ), request).select_related('truck', 'driver')
        if truck_id:
            qs = qs.filter(truck_id=truck_id)
        qs = qs.order_by('loading_time')
        headers = ['Waybill', 'Truck', 'Driver', 'Route',
                   'Revenue (GH₵)', 'Fuel Cost (GH₵)', 'Spare Parts (GH₵)', 'Net Profit (GH₵)']
        rows = []
        for t in qs:
            rows.append([
                t.waybill_no, t.truck.truck_number, t.driver.name,
                f'{t.origin} → {t.destination}',
                _fmt(t.trip_revenue), _fmt(t.fuel_cost),
                _fmt(t.spare_parts_cost), _fmt(t.net_profit),
            ])
        summary = {
            'Total Revenue (GH₵)':     round(sum(float(r[4]) for r in rows), 2),
            'Total Fuel Cost (GH₵)':   round(sum(float(r[5]) for r in rows), 2),
            'Total Spare Parts (GH₵)': round(sum(float(r[6]) for r in rows), 2),
            'Net Profit (GH₵)':        round(sum(float(r[7]) for r in rows), 2),
        }
        return _respond(request, headers, rows, summary, 'Trip P&L')


class TruckWiseSummaryView(BranchFilterMixin, APIView):
    def get(self, request):
        date_from, date_to = _parse_dates(request)
        truck_id = request.query_params.get('truck')
        qs = _apply_branch(Truck.objects.all(), request)
        if truck_id:
            qs = qs.filter(id=truck_id)
        headers = [
            'Truck', 'Model', 'Status', 'Trips',
            'Trip Revenue (GH₵)', 'Other Revenue (GH₵)', 'Total Revenue (GH₵)',
            'Fuel Cost (GH₵)', 'Maintenance (GH₵)', 'Tyre (GH₵)',
            'Spare Parts (GH₵)', 'Driver Wage (GH₵)', 'Toll (GH₵)',
            'Other Exp. (GH₵)', 'Total Expenditure (GH₵)', 'Net Profit (GH₵)',
        ]
        rows = []
        for truck in qs:
            trips = Trip.objects.filter(
                truck=truck,
                loading_time__date__gte=date_from,
                loading_time__date__lte=date_to,
            )
            trip_count = trips.count()
            trip_rev   = _fmt(trips.aggregate(t=Sum('trip_revenue'))['t'])
            other_rev  = _fmt(Revenue.objects.filter(
                date__gte=date_from, date__lte=date_to, trip__truck=truck,
            ).exclude(source='TRIP_REVENUE').aggregate(t=Sum('amount'))['t'])
            total_rev  = round(trip_rev + other_rev, 2)
            exp_qs = Expenditure.objects.filter(truck=truck, date__gte=date_from, date__lte=date_to)
            def _exp_cat(cat):
                return _fmt(exp_qs.filter(category=cat).aggregate(t=Sum('amount'))['t'])
            fuel_exp  = _exp_cat(Expenditure.FUEL)
            maint_exp = _exp_cat(Expenditure.MAINTENANCE)
            tyre_exp  = _exp_cat(Expenditure.TYRE)
            spare_exp = _exp_cat(Expenditure.SPARE_PART)
            wage_exp  = _exp_cat(Expenditure.DRIVER_WAGE)
            toll_exp  = _exp_cat(Expenditure.TOLL)
            other_exp = round(_fmt(exp_qs.filter(
                category__in=[Expenditure.ADMIN, Expenditure.OTHER]
            ).aggregate(t=Sum('amount'))['t']), 2)
            total_exp = round(fuel_exp + maint_exp + tyre_exp + spare_exp + wage_exp + toll_exp + other_exp, 2)
            net = round(total_rev - total_exp, 2)
            if trip_count == 0 and total_rev == 0 and total_exp == 0 and not truck_id:
                continue
            rows.append([
                truck.truck_number, truck.model, truck.get_status_display(), trip_count,
                trip_rev, other_rev, total_rev,
                fuel_exp, maint_exp, tyre_exp, spare_exp, wage_exp, toll_exp, other_exp,
                total_exp, net,
            ])
        summary = {
            'Total Revenue (GH₵)':     round(sum(float(r[6])  for r in rows), 2),
            'Total Expenditure (GH₵)': round(sum(float(r[14]) for r in rows), 2),
            'Net Profit (GH₵)':        round(sum(float(r[15]) for r in rows), 2),
        }
        return _respond(request, headers, rows, summary, 'Truck-wise Summary')


class StockReportView(BranchFilterMixin, APIView):
    def get(self, request):
        user = request.user
        if getattr(user, 'role', None) == 'SUPER_ADMIN':
            param = request.query_params.get('branch_id')
            if param:
                try:
                    branch_id = int(param)
                    from apps.users.models import Branch as BranchModel
                    branch = BranchModel.objects.filter(pk=branch_id).first()
                except (ValueError, TypeError):
                    branch_id = None
                    branch = None
            else:
                branch_id = None
                branch = None
        else:
            branch_id = user.branch_id
            branch = user.branch if branch_id else None

        if branch_id is not None:
            items = Item.objects.filter(
                ledger_entries__branch_id=branch_id
            ).distinct().order_by('item_type', 'name')
        elif getattr(user, 'role', None) == 'SUPER_ADMIN':
            items = Item.objects.filter(ledger_entries__isnull=False).distinct().order_by('item_type', 'name')
        else:
            items = Item.objects.none()
        headers = ['Item', 'Type', 'Unit', 'Qty in Stock', 'Stock Value (GH₵)', 'Reorder Level']
        rows = []
        total_value = Decimal('0')
        for item in items:
            qty   = item.available_qty(branch=branch)
            value = item.available_value(branch=branch)
            total_value += value
            rows.append([item.name, item.get_item_type_display(), item.unit,
                         _fmt(qty), _fmt(value), _fmt(item.reorder_level)])
        summary = {
            'Total Items':       len(rows),
            'Total Value (GH₵)': _fmt(total_value),
            'Low Stock Items':   sum(1 for r in rows if float(r[3]) <= float(r[5]) and float(r[5]) > 0),
        }
        return _respond(request, headers, rows, summary, 'Stock Report')


class InvoiceReportView(BranchFilterMixin, APIView):
    def get(self, request):
        date_from, date_to = _parse_dates(request)
        invoices = _apply_branch(Invoice.objects.filter(
            invoice_date__gte=date_from, invoice_date__lte=date_to
        ), request).order_by('invoice_date')
        headers = ['Invoice #', 'Client', 'Date', 'Status', 'Subtotal (GH₵)',
                   'VAT (GH₵)', 'Total (GH₵)', 'Paid (GH₵)', 'Balance (GH₵)']
        rows = []
        for inv in invoices:
            rows.append([
                inv.invoice_number, inv.client_name, str(inv.invoice_date),
                inv.get_status_display(),
                _fmt(inv.subtotal), _fmt(inv.vat_amount), _fmt(inv.total_amount),
                _fmt(inv.paid_amount), _fmt(inv.balance_due),
            ])
        agg = invoices.aggregate(
            total=Sum('total_amount'), paid=Sum('paid_amount'),
            balance=Sum('balance_due'), vat=Sum('vat_amount'),
        )
        summary = {
            'Total Invoiced (GH₵)': _fmt(agg['total']),
            'Total Paid (GH₵)':     _fmt(agg['paid']),
            'Balance Due (GH₵)':    _fmt(agg['balance']),
            'Total VAT (GH₵)':      _fmt(agg['vat']),
        }
        return _respond(request, headers, rows, summary, 'Invoice Report')


class SparePartsReportView(BranchFilterMixin, APIView):
    def get(self, request):
        date_from, date_to = _parse_dates(request)
        ledger = _apply_branch(
            StockLedger.objects.filter(
                item__item_type='SPARE_PART',
                created_at__date__gte=date_from,
                created_at__date__lte=date_to,
            ),
            request,
        ).select_related('item', 'location').order_by('created_at')
        headers = ['Date', 'Item', 'Transaction', 'Qty', 'Unit Cost (GH₵)', 'Total (GH₵)', 'Location', 'Reference']
        rows = []
        for entry in ledger:
            rows.append([
                str(entry.created_at.date()), entry.item.name,
                entry.get_transaction_type_display(),
                _fmt(entry.quantity), _fmt(entry.unit_price), _fmt(entry.final_amount),
                entry.location.name if entry.location else '—',
                entry.reference_type or '',
            ])
        agg = ledger.aggregate(total=Sum('final_amount'))
        summary = {'Total Value (GH₵)': _fmt(agg['total'])}
        return _respond(request, headers, rows, summary, 'Spare Parts Report')


class MaintenanceReportView(BranchFilterMixin, APIView):
    def get(self, request):
        date_from, date_to = _parse_dates(request)
        logs = _apply_branch(MaintenanceLog.objects.filter(
            service_date__gte=date_from, service_date__lte=date_to
        ), request).select_related('truck', 'mechanic').order_by('service_date')
        headers = ['Date', 'Truck', 'Type', 'Description', 'Mechanic',
                   'Labour Cost (GH₵)', 'Parts Cost (GH₵)', 'Total (GH₵)', 'Status']
        rows = []
        for log in logs:
            rows.append([
                str(log.service_date), log.truck.truck_number,
                log.get_maintenance_type_display(), log.description or '',
                log.mechanic.name if log.mechanic else '—',
                _fmt(log.labour_cost), _fmt(log.parts_cost), _fmt(log.total_cost),
                log.get_status_display(),
            ])
        agg = logs.aggregate(
            total_labour=Sum('labour_cost'),
            total_parts=Sum('parts_cost'),
            total_cost=Sum('total_cost'),
        )
        summary = {
            'Total Records':      len(rows),
            'Total Labour (GH₵)': _fmt(agg['total_labour']),
            'Total Parts (GH₵)':  _fmt(agg['total_parts']),
            'Total Cost (GH₵)':   _fmt(agg['total_cost']),
        }
        return _respond(request, headers, rows, summary, 'Maintenance Report')


class VATReportView(BranchFilterMixin, APIView):
    def get(self, request):
        date_from, date_to = _parse_dates(request)
        invoices = _apply_branch(Invoice.objects.filter(
            invoice_date__gte=date_from, invoice_date__lte=date_to,
            vat_applicable=True,
        ), request).order_by('invoice_date')
        headers = ['Invoice #', 'Client', 'Date', 'Subtotal (GH₵)', 'VAT %', 'VAT Amount (GH₵)', 'Total (GH₵)']
        rows = []
        for inv in invoices:
            rows.append([
                inv.invoice_number, inv.client_name, str(inv.invoice_date),
                _fmt(inv.subtotal), _fmt(inv.vat_percentage),
                _fmt(inv.vat_amount), _fmt(inv.total_amount),
            ])
        total_vat = invoices.aggregate(t=Sum('vat_amount'))['t'] or Decimal('0')
        summary   = {'Total VAT Collected (GH₵)': _fmt(total_vat)}
        return _respond(request, headers, rows, summary, 'VAT Report')


class TyreReportView(BranchFilterMixin, APIView):
    def get(self, request):
        tyres = _apply_branch(Tyre.objects.all(), request).prefetch_related('assignments__truck').order_by('status', 'serial_number')
        headers = ['Serial #', 'Brand', 'Model', 'Size', 'Status',
                   'Unit Cost (GH₵)', 'Truck Fitted', 'Position', 'KM Used']
        rows = []
        for tyre in tyres:
            asgn = tyre.current_assignment
            rows.append([
                tyre.serial_number, tyre.brand, tyre.model, tyre.size,
                tyre.get_status_display(), _fmt(tyre.unit_cost),
                asgn.truck.truck_number if asgn else '—',
                asgn.position          if asgn else '—',
                _fmt(asgn.km_used)     if asgn and asgn.km_used else '—',
            ])
        summary = {
            'Total Tyres': len(rows),
            'Fitted':      sum(1 for t in tyres if t.status == Tyre.FITTED),
            'In Store':    sum(1 for t in tyres if t.status == Tyre.STORE),
            'Condemned':   sum(1 for t in tyres if t.status == Tyre.CONDEMNED),
        }
        return _respond(request, headers, rows, summary, 'Tyre Report')


class LubricantReportView(BranchFilterMixin, APIView):
    def get(self, request):
        date_from, date_to = _parse_dates(request)
        ledger = _apply_branch(
            StockLedger.objects.filter(
                item__item_type='LUBRICANT',
                created_at__date__gte=date_from,
                created_at__date__lte=date_to,
            ),
            request,
        ).select_related('item', 'location').order_by('created_at')
        headers = ['Date', 'Item', 'Transaction', 'Qty', 'Unit', 'Unit Cost (GH₵)', 'Total (GH₵)', 'Location']
        rows = []
        for entry in ledger:
            rows.append([
                str(entry.created_at.date()), entry.item.name,
                entry.get_transaction_type_display(),
                _fmt(entry.quantity), entry.item.unit,
                _fmt(entry.unit_price), _fmt(entry.final_amount),
                entry.location.name if entry.location else '—',
            ])
        agg = ledger.aggregate(total=Sum('final_amount'))
        summary = {'Total Value (GH₵)': _fmt(agg['total'])}
        return _respond(request, headers, rows, summary, 'Lubricant Report')


class CleanupOrphanedRevenueView(APIView):
    def post(self, request):
        deleted, _ = Revenue.objects.filter(
            trip__isnull=False, trip__status=Trip.CANCELLED
        ).delete()
        return Response({'deleted': deleted})


class PurgePhantomTripsView(APIView):
    def post(self, request):
        cutoff = timezone.now().date() - timedelta(days=90)
        qs = Trip.objects.filter(
            status=Trip.PLANNED,
            loading_time__date__lt=cutoff,
            revenue_entries__isnull=True,
        )
        count = qs.count()
        qs.delete()
        return Response({'purged': count})
